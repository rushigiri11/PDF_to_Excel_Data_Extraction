import os
import fitz  # PyMuPDF
import re
import openpyxl

# Function to extract specific data from a PDF text
def extract_data_from_pdf(pdf_path):
    # Open the PDF file
    pdf_document = fitz.open(pdf_path)

    # Define the regex patterns for the data to be extracted
    patterns = {
        "Beneficiary Name": r"3\.\nBeneficiary Name\n(.*?)\n",
        "Application No": r"2\.\nApplication No\n(.*?)\n",
        "Location": r"Location\n(.*?)\n",
        "Motor Capacity (HP/KW)": r"Motor Capacity \(HP/KW\)\s*:\s*(.*?)(?=\n|$)",
        "Controller Id / IMEI No": r"Controller Id / IMEI No\s*:\s*(.*?)(?=\n|$)",
        "Pump Id": r"Pump Id\s*:\s*(.*?)(?=\n|$)",
        "PV Panel Serial No.": r"PV Panel Serial No\.\s*:\s*((?:.*\n?)+?)\s*(?=\n\S+:|\Z)",
    }

    data_list = []

    for page_num in range(len(pdf_document)):
        page = pdf_document.load_page(page_num)
        text = page.get_text()

        # Extract data using the defined patterns
        extracted_data = {}
        for key, pattern in patterns.items():
            match = re.search(pattern, text)
            if match:
                value = match.group(1).strip().replace('\n', ' ')
                if key == "PV Panel Serial No.":
                    # Remove anything after 'Yes' in PV Panel Serial No.
                    value = re.split(r'\s*Yes\s*', value)[0].strip()
                    value = value.replace(' ', '').strip()

                if key == "Pump Id":
                    # Remove 'Yes' from Pump Id
                    value = value.replace('Yes', '').strip()
                extracted_data[key] = value
            else:
                extracted_data[key] = None

        # Split Location into District, Taluka, and Village
        location = extracted_data.get("Location", "")
        if location:
            location_parts = location.split(',')
            extracted_data["District"] = location_parts[0].strip() if len(location_parts) > 0 else None
            extracted_data["Taluka"] = location_parts[1].strip() if len(location_parts) > 1 else None
            extracted_data["Village"] = location_parts[2].strip() if len(location_parts) > 2 else None
        else:
            extracted_data["District"] = None
            extracted_data["Taluka"] = None
            extracted_data["Village"] = None

        # Remove the original Location key as it's now split
        if "Location" in extracted_data:
            del extracted_data["Location"]

        data_list.append(extracted_data)

    pdf_document.close()

    return data_list

# Function to write data to an Excel file
def write_data_to_excel(data_list, excel_path):
    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Define the column headings
    columns = ["Beneficiary Name", "Application No", "District", "Taluka", "Village",
               "Motor Capacity (HP/KW)", "Controller Id / IMEI No",
               "Pump Id", "PV Panel Serial No."]

    # Write the column headings to the first row
    for col_num, column_title in enumerate(columns, 1):
        sheet.cell(row=1, column=col_num, value=column_title)

    # Write the data to the worksheet
    for row_num, data in enumerate(data_list, 2):
        for col_num, column_title in enumerate(columns, 1):
            sheet.cell(row=row_num, column=col_num, value=data.get(column_title, ''))

    # Save the workbook to the specified path
    workbook.save(excel_path)

# Main function to extract data from all PDFs in a folder and write to Excel
def main():
    pdf_folder = "pdfdata"
    excel_path = "output_data.xlsx"

    # Initialize an empty list to store data from all PDFs
    all_data = []

    # Iterate over all files in the PDF folder
    for filename in os.listdir(pdf_folder):
        if filename.endswith(".pdf"):
            pdf_path = os.path.join(pdf_folder, filename)
            # Extract data from the PDF
            data_list = extract_data_from_pdf(pdf_path)
            # Extend the all_data list with the data from the current PDF
            all_data.extend(data_list)

    # Write the extracted data from all PDFs to an Excel file
    write_data_to_excel(all_data, excel_path)

    print(f"Data successfully extracted from all PDFs and saved to {excel_path}")

# Run the main function
if __name__ == "__main__":
    main()
